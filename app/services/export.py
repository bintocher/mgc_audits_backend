from typing import List, Dict, Any
from datetime import datetime
from io import BytesIO
import json
import zipfile
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


def export_findings_to_excel(data: List[Dict]) -> BytesIO:
    """
    Экспортировать отчет по несоответствиям в Excel.
    
    Args:
        data: Список словарей с данными несоответствий
    
    Returns:
        BytesIO: Excel файл в памяти
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Несоответствия"
    
    headers = [
        "№", "№ Аудита", "Аудит", "Название", "Тип", "Процесс",
        "Статус", "Исполнитель", "Утверждающий", "Дедлайн", "Дата закрытия", "Дата создания"
    ]
    
    ws.append(headers)
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    for row_data in data:
        ws.append([
            row_data.get("finding_number"),
            row_data.get("audit_number"),
            row_data.get("audit_title"),
            row_data.get("title"),
            row_data.get("finding_type"),
            row_data.get("process_name"),
            row_data.get("status_name"),
            row_data.get("resolver_name"),
            row_data.get("approver_name"),
            row_data.get("deadline"),
            row_data.get("closing_date"),
            row_data.get("created_at")
        ])
    
    for col in range(1, len(headers) + 1):
        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].width = 20
    
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output


def export_process_to_excel(data: List[Dict]) -> BytesIO:
    """
    Экспортировать отчет по процессам в Excel.
    
    Args:
        data: Список словарей с данными по процессам
    
    Returns:
        BytesIO: Excel файл в памяти
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "По процессам"
    
    headers = [
        "Процесс", "Всего", "CAR1", "CAR2", "OFI", "Закрыто", "Просрочено"
    ]
    
    ws.append(headers)
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    for row_data in data:
        ws.append([
            row_data.get("process_name"),
            row_data.get("total_findings"),
            row_data.get("car1_count"),
            row_data.get("car2_count"),
            row_data.get("ofi_count"),
            row_data.get("closed_count"),
            row_data.get("overdue_count")
        ])
    
    for col in range(1, len(headers) + 1):
        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].width = 20
    
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(horizontal="left", vertical="center")
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output


def export_solver_to_excel(data: List[Dict]) -> BytesIO:
    """
    Экспортировать отчет по исполнителям в Excel.
    
    Args:
        data: Список словарей с данными по исполнителям
    
    Returns:
        BytesIO: Excel файл в памяти
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "По исполнителям"
    
    headers = [
        "Исполнитель", "Всего", "Активных", "Просрочено", "Закрыто"
    ]
    
    ws.append(headers)
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    for row_data in data:
        ws.append([
            row_data.get("solver_name"),
            row_data.get("total_findings"),
            row_data.get("active_findings"),
            row_data.get("overdue_findings"),
            row_data.get("closed_findings")
        ])
    
    for col in range(1, len(headers) + 1):
        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].width = 20
    
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(horizontal="left", vertical="center")
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output


def export_audit_to_zip(
    audit_data: Dict[str, Any],
    findings_data: List[Dict[str, Any]],
    attachments_data: List[Dict[str, Any]],
    history_data: List[Dict[str, Any]]
) -> BytesIO:
    """
    Экспортировать аудит со всеми связанными данными в ZIP архив.
    
    Args:
        audit_data: Данные аудита
        findings_data: Список связанных несоответствий
        attachments_data: Список вложений
        history_data: История изменений
    
    Returns:
        BytesIO: ZIP архив в памяти
    """
    zip_buffer = BytesIO()
    
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        audit_json = json.dumps(audit_data, ensure_ascii=False, indent=2, default=str)
        zip_file.writestr("audit.json", audit_json.encode('utf-8'))
        
        if findings_data:
            findings_json = json.dumps(findings_data, ensure_ascii=False, indent=2, default=str)
            zip_file.writestr("findings.json", findings_json.encode('utf-8'))
        
        if history_data:
            history_json = json.dumps(history_data, ensure_ascii=False, indent=2, default=str)
            zip_file.writestr("history.json", history_json.encode('utf-8'))
        
        if attachments_data:
            attachments_json = json.dumps(attachments_data, ensure_ascii=False, indent=2, default=str)
            zip_file.writestr("attachments.json", attachments_json.encode('utf-8'))
    
    zip_buffer.seek(0)
    return zip_buffer

